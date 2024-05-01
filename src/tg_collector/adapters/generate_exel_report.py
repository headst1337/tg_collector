from typing import List
from io import BytesIO
from datetime import datetime
import tempfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from tg_collector.application.models import Post
from tg_collector.application.protocols import GenerateDateReport
from tg_collector.adapters.sqlalchemy_db.repository import PostRepository


class GenerateExelReport(GenerateDateReport):
    def __init__(self) -> None:
        self.posrt_repository = PostRepository()

    def genetare_report(self, dates: List[str]) -> str:
        data = self.__prepare_data(dates)
        file_content =  self.__build_file_conent(data)
        return self.__build_file(file_content)
    
    def __prepare_data(self, dates: List[str]) -> List[Post]:
        data = []
        for date in dates:
            date = datetime.strptime(date, "%d-%m-%Y").date()
            data.extend(self.posrt_repository.get_by_date(date))
        return data
    
    def __build_file(self, file_content: BytesIO) -> str:
        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            temp_file.write(file_content.getvalue())
            return temp_file.name
    
    def __build_file_conent(self, data: List[Post]) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Ссылка на пост'
        sheet['C1'] = 'Ссылка на группу'
        sheet['D1'] = 'Содержимое'
        sheet['E1'] = 'Платежные данные'
        sheet['F1'] = 'Дата'

        bold_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = bold_font

        for row, post in enumerate(data, start=2):
            sheet.cell(row=row, column=1, value=post.id)
            sheet.cell(row=row, column=2, value=post.post_link)
            sheet.cell(row=row, column=3, value=post.group_link)
            sheet.cell(row=row, column=4, value=post.body)

            payment_cell = sheet.cell(
                row=row,
                column=5,
                value=post.payment_data,
            )
            payment_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True,
            )

            sheet.cell(row=row, column=6, value=post.timestamp)

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[
                get_column_letter(column[0].column)
            ].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
